import openpyxl
from openpyxl.styles import Font 
import os
import math
from tkinter import *
import tkinter.messagebox


def cmpSame(s1,s2):
    # a = difflib.SequenceMatcher(None, s1, s2).quick_ratio()
    # if(a>0.85):
    #     return True
    # else:
    #     return False
    if(len(s1)==len(s2)):
        for i in range(0,len(s1)-1):
            if(s1[i]!=s2[i]):
                return False
        return True
    elif(math.fabs(len(s1)-len(s2))<=1):
        for i in range(0,min(len(s1),len(s2))):
            if(s1[i]!=s2[i]):
                return False
        return True
    else:
        return False

def mainfunc(s):
    # wb = openpyxl.load_workbook(r'C:\Users\user\Desktop\workspace\python\excel\data.xlsx')
    wb = openpyxl.load_workbook(s)

    sheet = wb.worksheets[0]

    valName = []
    valNum = []
    cmp = {}

    for rowStr in range(2,sheet.max_row,1):
        valName.append(sheet.cell(rowStr, 1 ).value)
        # print(sheet.cell(rowStr, 1 ).value)
    for rowNum in range(2,sheet.max_row):
        valNum.append(sheet.cell(rowNum, 2).value)
        # print(sheet.cell(rowNum, 2).value)
    for i in range(0, len(valName)):
        cmp[valName[i]] = valNum[i]
    # for i in range(0, len(valName)):
    #     print(cmp[valName[i]]) 

    valName.append('')
    valNum.append(0)

    cnt = 2
    iscon = False
    sheet.cell(1,6).value = '新的列標籤'
    sheet.cell(1,7).value = "數量"
    for i in range(0,len(valName)-1):
        if(iscon):
            iscon = False
            continue
        if(cmpSame(valName[i],valName[i+1]) == True):
            sheet.cell(cnt,6).value =  valName[i]+'+'+valName[i+1]
            sheet.cell(cnt,7).value = valNum[i]+valNum[i+1]
            iscon = True
        else:
            sheet.cell(cnt,6).value = valName[i]
            sheet.cell(cnt,7).value = valNum[i]
        cnt+=1
        
    # wb.save(r'C:\Users\user\Desktop\workspace\python\excel\data1.xlsx')
    wb.save(s)
    print("Done")





#圖形化介面處理

def inspect(val):
    filena, extension = os.path.splitext(val)
    if(os.path.exists(val) and extension == '.xlsx'):
        mainfunc(val)
        anwser = tkinter.messagebox.askokcancel('Success',val+'已完成處理')
        return False
    else:
        anwser = tkinter.messagebox.askokcancel('Warning',val+'並非正確的檔案位置，請輸入正確檔案位置及檔名')
        return True


root = Tk()
root.geometry('460x240')
root.title("Excel")


lab1 = Label(root,text='請輸入excel檔案位置(絕對路徑)')
lab1.place(relx=0.3,rely=0.1)
lab2 = Label(root,text='範例:  C:\\Users\\excel\\data.xlsx')
lab2.place(relx=0.3,rely=0.2)

inp1 = Entry(root)
inp1.place(relx=0.15,rely=0.4,relwidth=0.7,relheight=0.1)

btn = Button(root,text='執行',command=lambda: inspect(inp1.get()))
btn.place(relx = 0.4,rely=0.6,relwidth=0.2,relheight=0.1)

root.mainloop()

#--------------------------------------

# while(1):
#     print("請輸入excel檔案位置(絕對路徑)   輸入E結束程式:"+'\n'+"範例:  C:\\Users\\excel\\data.xlsx")
#     val = input()
#     if(val=='E'):
#         break
#     if(val == "cls"):
#         os.system("cls")
#         continue
#     root, extension = os.path.splitext(val)
#     if(os.path.exists(val) and extension == '.xlsx'):
#         mainfunc(val)
#     else:
#         print(val+"並非正確的檔案位置，請輸入正確檔案位置及檔名")
#     print("")
